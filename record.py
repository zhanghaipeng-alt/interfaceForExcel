'''
    使用mitmproxy 录制接口信息，写入到Excel中
'''
from openpyxl import Workbook
from mitmproxy import ctx
import xlrd
import json

class Counter:
    def __init__(self):
        self.num = 2
        self.wb = Workbook()
        # 创建一个sheet页
        self.ws = self.wb.create_sheet()
        title = ['URL', '请求方法类型', '请求头']

        for i, v in enumerate(title):
            self.ws.cell(row=1, column=i+1, value=v)
        self.wb.save('./demo1.xlsx')

    # 过滤某接口信息
    def request(self, flow):
        if flow.request.url.startswith('http://140.143.171.176:9000/'):
            data = [flow.request.url, flow.request.method, json.dumps(flow.request.headers)]
            for i, v in enumerate(data):
                self.ws.cell(row=self.num, column=i+1, value=v)
            self.wb.save('./demo1.xlsx')
            self.num += 1

addons = [
    Counter()
    ]