import requests
from core.util import *
from core.client import *
from openpyxl import Workbook
import xlrd
import jsonpath

DATA = get_dic_global('全局变量')
TEMPLES = get_dic_excel('接口模板')
cases = get_dic_excel('测试用例')
demo_list = get_dic_excel('demo')
info = []
def aa(key, value):
    dic = {}
    dic['key'] = value
    return dic

def res_test(url, headers,data):
    res = requests.get(headers=headers,url=url,params=data)
    return  res.headers, res.url, res.content, res.cookies,res.status_code, res.elapsed.total_seconds()*1000

def testlogic():
    list1 = [1,3,6,7,8,2]

    for i in list1:
        if i == 7:
            print('找到了')
            break
    else:
        print('没有')

# 写入到Excel
def enter_excel():
    wb = Workbook()
    title = ['URL', '请求方法类型', '请求头']
    sheet = wb.create_sheet('测试模板')
    for i, v in enumerate(title):
        sheet.cell(1, i+1, v)

    wb.save('./demo.xlsx')

# 读取Excel
def read_excel():
    d = []
    wb = xlrd.open_workbook('/Users/zhanghaipeng/PycharmProjects/InterfaceTestForExcel/demo.xlsx')
    sheet = wb.sheet_by_name('测试模板')
    for i in range(sheet.nrows):
       print(sheet.row_values(i, i))




if __name__ == '__main__':
    read_excel()